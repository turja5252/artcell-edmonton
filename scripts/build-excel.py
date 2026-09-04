#!/usr/bin/env python3
"""Build the Artcell Edmonton shared Excel workbook."""

from copy import copy
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import DataBarRule
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.marker import DataPoint as DP

ROOT = Path(__file__).resolve().parents[1]

SPONSORS = [
    ("Dr Jasmeen", "", False, ""),
    ("Millet Pharmacy", "Novel", False, ""),
    ("Desjardin", "Nick", False, ""),
    ("Bishwajit Barua", "Nick", False, ""),
    ("MR Engineering", "Fahim", False, ""),
    ("Chandu Talukdar - CN Tax", "", False, ""),
    ("Mahabub Mollah", "Sathi", False, ""),
    ("Razib Chowdhury", "Khaled", False, ""),
    ("Gurpinder (Mortgage)", "Sathi", False, ""),
    ("Dominion Landing", "Fahim", False, ""),
    ("Raiyan's Company", "Wade Engineering", False, ""),
    ("Kumon and Abacus", "Mrs Mizanur Shoma Apa", False, ""),
    ("TK Auto", "Shenin", False, ""),
    ("Elite Integrity Service", "Tanzim", False, ""),
    ("Saem Daily Bazaar", "Tanzim", False, ""),
    ("Mohsin Realton", "Tanzim", False, ""),
    ("Iron Site", "Tahiat", False, ""),
    ("Canadian Butcher", "Khaled Bhai", False, ""),
    ("SVF", "Khaled Bhai", False, ""),
    ("Top Donair", "Khaled Bhai", False, ""),
    ("Realtor Duke Bhai", "Khaled Bhai", False, ""),
    ("Physio", "Khaled Bhai", False, ""),
    ("Madira", "Sathi", False, ""),
    ("Sudipto", "Sathi", False, ""),
    ("Shatkahon", "Sathi", False, ""),
    ("Sheila", "Tanvir/Tajul Bhai", False, ""),
    ("Ketek", "Raiyan", False, ""),
    ("Tanvir Bhai", "Shuddho", False, ""),
    ("Donair & Poutine", "Fahim", False, ""),
    ("Mizan Bhai", "Fahim", False, ""),
    ("Dominion Laundry", "Fahim", False, ""),
    ("New conl", "Shenin", False, ""),
]

SETLIST = [
    ("Dukkho Bilash Outro", "8:15", "https://www.youtube.com/watch?v=xSue3Ckaoos&t=495s"),
    ("Onno Shomoy Intro", "10:52", "https://www.youtube.com/watch?v=xSue3Ckaoos&t=652s"),
    ("Onno Shomoy Instrumental Section", "1:42", "https://www.youtube.com/watch?v=v6SWr2UeYs4&t=102s"),
    ("Onno Shomoy Outro", "4:22", "https://www.youtube.com/watch?v=v6SWr2UeYs4&t=262s"),
    ("Dhushor Shomoy Final Chorus", "6:18", "https://www.youtube.com/watch?v=eo4Zj-7Ex4o&t=378s"),
    ("Rahur Ghrash Chorus", "4:52", "https://www.youtube.com/watch?v=7FupQWRuqSc&t=292s"),
    ("Rahur Grash Outro Instrumental", "6:39", "https://www.youtube.com/watch?v=7FupQWRuqSc&t=399s"),
    ("Bhul Jonmo Chorus", "2:02", "https://www.youtube.com/watch?v=z47nnlHkMIc&t=122s"),
    ("Utshober Utshahe", "4:15", "https://www.youtube.com/watch?v=RpEFOwsHEmE&t=255s"),
    ("Chile Kothar Sepai Solo", "3:11", "https://www.youtube.com/watch?v=xI_Fa-wpGgw&t=191s"),
    ("Chile Kothar Shepai Instrumental", "4:31", "https://www.youtube.com/watch?v=ZI0TYzNuSq4&t=271s"),
]

NAVY = "1C140E"
GOLD = "C9A227"
CREAM = "F7F1E6"
YELLOW = "FFE8A3"
GRAY = "EEE6D8"
GREEN = "1F7A4D"
RED = "8B2E2E"
WHITE = "FFFFFF"
INK = "2A2118"

thin = Border(
    left=Side(style="thin", color="D4C4A8"),
    right=Side(style="thin", color="D4C4A8"),
    top=Side(style="thin", color="D4C4A8"),
    bottom=Side(style="thin", color="D4C4A8"),
)


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def font(name="Calibri", size=11, bold=False, color=INK) -> Font:
    return Font(name=name, size=size, bold=bold, color=color)


def apply_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_header_row(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.fill = fill(NAVY)
        cell.font = font(size=11, bold=True, color=GOLD)
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def add_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)


def build() -> Path:
    wb = Workbook()

    dash = wb.active
    dash.title = "Dashboard"

    sponsors = wb.create_sheet("Sponsors")
    attendance = wb.create_sheet("Attendance")
    setlist = wb.create_sheet("Setlist")
    howto = wb.create_sheet("How to share")

    # --- Sponsors ---
    sponsor_headers = [
        "Company",
        "Assigned",
        "Done",
        "Outcome",
        "Committed",
        "Received",
        "Still owing",
    ]
    sponsors.append(sponsor_headers)
    style_header_row(sponsors, 1, len(sponsor_headers))
    sponsors.row_dimensions[1].height = 22
    sponsors.freeze_panes = "A2"
    sponsors.auto_filter.ref = f"A1:G{len(SPONSORS) + 1}"

    for i, (company, assigned, done, outcome) in enumerate(SPONSORS, start=2):
        sponsors.cell(i, 1, company)
        sponsors.cell(i, 2, assigned)
        done_cell = sponsors.cell(i, 3, "No")
        done_cell.alignment = Alignment(horizontal="center")
        sponsors.cell(i, 4, outcome)
        committed = sponsors.cell(i, 5, 0)
        received = sponsors.cell(i, 6, 0)
        owing = sponsors.cell(i, 7, f"=MAX(0,E{i}-F{i})")
        committed.number_format = '"$"#,##0'
        received.number_format = '"$"#,##0'
        owing.number_format = '"$"#,##0'
        owing.fill = fill(GRAY)
        committed.fill = fill(YELLOW)
        received.fill = fill(YELLOW)
        done_cell.fill = fill(YELLOW)

    last_s = len(SPONSORS) + 1
    add_table(sponsors, "Sponsors", f"A1:G{last_s}")

    done_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    done_dv.error = "Pick Yes or No"
    done_dv.errorTitle = "Done"
    sponsors.add_data_validation(done_dv)
    done_dv.add(f"C2:C{last_s}")

    outcome_dv = DataValidation(
        type="list",
        formula1='"Waiting for reply,Meeting booked,Confirmed sponsor,In-kind support,Declined,Can\'t reach"',
        allow_blank=True,
    )
    sponsors.add_data_validation(outcome_dv)
    outcome_dv.add(f"D2:D{last_s}")

    apply_widths(sponsors, {"A": 32, "B": 24, "C": 10, "D": 24, "E": 14, "F": 14, "G": 14})
    sponsors.sheet_properties.pageSetUpPr.fitToPage = True
    sponsors.page_setup.orientation = "landscape"
    sponsors.page_setup.fitToWidth = 1
    sponsors.page_setup.fitToHeight = 0
    sponsors.oddHeader.left.text = "Artcell Edmonton — Sponsors"
    sponsors.sheet_view.showGridLines = False

    # --- Attendance ---
    att_headers = ["Name or group", "Assigned", "Status", "Seats", "Notes"]
    attendance.append(att_headers)
    style_header_row(attendance, 1, len(att_headers))
    attendance.row_dimensions[1].height = 22
    attendance.freeze_panes = "A2"

    blank_rows = 25
    for i in range(2, blank_rows + 2):
        attendance.cell(i, 1, "")
        attendance.cell(i, 2, "")
        status = attendance.cell(i, 3, "Not reached")
        seats = attendance.cell(i, 4, 1)
        attendance.cell(i, 5, "")
        status.fill = fill(YELLOW)
        seats.fill = fill(YELLOW)
        attendance.cell(i, 1).fill = fill(YELLOW)
        attendance.cell(i, 2).fill = fill(YELLOW)
        attendance.cell(i, 5).fill = fill(YELLOW)

    last_a = blank_rows + 1
    add_table(attendance, "Attendance", f"A1:E{last_a}")

    status_dv = DataValidation(
        type="list",
        formula1='"Not reached,Reached,Maybe,Confirmed,Declined"',
        allow_blank=False,
    )
    attendance.add_data_validation(status_dv)
    status_dv.add(f"C2:C{last_a}")

    apply_widths(attendance, {"A": 32, "B": 22, "C": 16, "D": 10, "E": 36})
    attendance.sheet_view.showGridLines = False
    attendance.oddHeader.left.text = "Artcell Edmonton — Attendance outreach"

    # --- Setlist ---
    set_headers = ["#", "Cue", "Timestamp", "YouTube"]
    setlist.append(set_headers)
    style_header_row(setlist, 1, len(set_headers))
    for i, (label, stamp, url) in enumerate(SETLIST, start=2):
        setlist.cell(i, 1, i - 1)
        setlist.cell(i, 2, label)
        setlist.cell(i, 3, stamp)
        cell = setlist.cell(i, 4, url)
        cell.hyperlink = url
        cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")
    add_table(setlist, "Setlist", f"A1:D{len(SETLIST) + 1}")
    apply_widths(setlist, {"A": 6, "B": 40, "C": 14, "D": 62})
    setlist.sheet_view.showGridLines = False

    # --- Dashboard ---
    dash.sheet_view.showGridLines = False
    dash.page_setup.orientation = "portrait"
    dash.page_setup.fitToPage = True
    dash.page_setup.fitToWidth = 1
    dash.page_setup.fitToHeight = 1
    dash.freeze_panes = "A8"
    apply_widths(dash, {"A": 3, "B": 22, "C": 18, "D": 3, "E": 22, "F": 18, "G": 3, "H": 22, "I": 16})
    dash.row_dimensions[1].height = 28
    dash.row_dimensions[2].height = 20
    dash.merge_cells("B1:F1")
    dash.merge_cells("B2:F2")
    title = dash["B1"]
    title.value = "ARTCELL  ·  EDMONTON SHOW"
    title.font = Font(name="Calibri", size=22, bold=True, color=NAVY)
    dash["B2"].value = "Type only in the yellow cells. Gray cells are running totals — they update themselves."
    dash["B2"].font = font(size=11, color="6B5B45")

    # Targets
    dash["B4"] = "MONEY TARGET (CAD)"
    dash["B4"].font = font(size=10, bold=True, color="6B5B45")
    dash["C4"] = 0
    dash["C4"].fill = fill(YELLOW)
    dash["C4"].number_format = '"$"#,##0'
    dash["C4"].font = font(size=16, bold=True)
    dash["C4"].border = thin

    dash["E4"] = "SEAT TARGET"
    dash["E4"].font = font(size=10, bold=True, color="6B5B45")
    dash["F4"] = 0
    dash["F4"].fill = fill(YELLOW)
    dash["F4"].number_format = "#,##0"
    dash["F4"].font = font(size=16, bold=True)
    dash["F4"].border = thin

    def card(cell, label, formula, money=False, seats=False):
        label_cell = dash[cell]
        # label above is set by caller using nearby cells
        value_cell = dash[cell]
        value_cell.value = formula
        value_cell.font = Font(name="Calibri", size=20, bold=True, color=NAVY)
        value_cell.fill = fill(GRAY)
        value_cell.border = thin
        value_cell.alignment = Alignment(horizontal="left", vertical="center")
        if money:
            value_cell.number_format = '"$"#,##0'
        if seats:
            value_cell.number_format = "#,##0"

    # Money block
    dash.merge_cells("B6:C6")
    dash["B6"] = "MONEY"
    dash["B6"].font = Font(name="Calibri", size=14, bold=True, color=GOLD)
    dash["B6"].fill = fill(NAVY)
    dash["C6"].fill = fill(NAVY)

    rows_money = [
        (7, "Committed", '=SUM(Sponsors[Committed])', True, False),
        (8, "Received", '=SUM(Sponsors[Received])', True, False),
        (9, "Remaining to target", '=MAX(0,C4-C7)', True, False),
        (10, "Still to collect", '=MAX(0,C7-C8)', True, False),
        (11, "Of target", '=IF(C4=0,"Set a target",C7/C4)', False, False),
        (12, "Companies pledged", '=COUNTIF(Sponsors[Committed],">0")', False, False),
    ]
    for row, label, formula, money, seats in rows_money:
        dash.cell(row, 2, label).font = font(size=11)
        dash.cell(row, 2).alignment = Alignment(vertical="center")
        card(f"C{row}", label, formula, money=money, seats=seats)
        dash.row_dimensions[row].height = 28
    dash["C11"].number_format = '0%'

    # Attendance block
    dash.merge_cells("E6:F6")
    dash["E6"] = "ATTENDANCE"
    dash["E6"].font = Font(name="Calibri", size=14, bold=True, color=GOLD)
    dash["E6"].fill = fill(NAVY)
    dash["F6"].fill = fill(NAVY)

    rows_att = [
        (7, "Confirmed seats", '=SUMIF(Attendance[Status],"Confirmed",Attendance[Seats])'),
        (8, "Maybe", '=SUMIF(Attendance[Status],"Maybe",Attendance[Seats])'),
        (9, "Remaining to fill", '=MAX(0,F4-F7)'),
        (10, "Not reached yet", '=SUMIF(Attendance[Status],"Not reached",Attendance[Seats])'),
        (11, "Of room", '=IF(F4=0,"Set a target",F7/F4)'),
        (12, "People on list", '=COUNTA(Attendance[Name or group])'),
    ]
    for row, label, formula in rows_att:
        dash.cell(row, 5, label).font = font(size=11)
        dash.cell(row, 5).alignment = Alignment(vertical="center")
        card(f"F{row}", label, formula, seats=True)
        dash.row_dimensions[row].height = 28
    dash["F11"].number_format = '0%'

    # Outreach block
    dash.merge_cells("B14:C14")
    dash["B14"] = "SPONSOR OUTREACH"
    dash["B14"].font = Font(name="Calibri", size=14, bold=True, color=GOLD)
    dash["B14"].fill = fill(NAVY)
    dash["C14"].fill = fill(NAVY)

    dash["B15"] = "Reached / done"
    dash["C15"] = '=COUNTIF(Sponsors[Done],"Yes")&" / "&COUNTA(Sponsors[Company])'
    dash["B16"] = "Still open"
    dash["C16"] = '=COUNTIF(Sponsors[Done],"No")'
    dash["B17"] = "Need an owner"
    dash["C17"] = '=COUNTIFS(Sponsors[Assigned],"",Sponsors[Done],"No")'
    for r in (15, 16, 17):
        dash.cell(r, 3).fill = fill(GRAY)
        dash.cell(r, 3).border = thin
        dash.cell(r, 3).font = Font(name="Calibri", size=16, bold=True, color=NAVY)

    dash.merge_cells("E14:F14")
    dash["E14"] = "WHAT TO TYPE"
    dash["E14"].font = Font(name="Calibri", size=14, bold=True, color=GOLD)
    dash["E14"].fill = fill(NAVY)
    dash["F14"].fill = fill(NAVY)
    dash.merge_cells("E15:F17")
    dash["E15"] = (
        "1. Set the two yellow targets above.\n"
        "2. On Sponsors, log $ they pledged and $ received.\n"
        "3. On Attendance, add names, seats, and status.\n"
        "4. This page updates itself. Don’t type over gray cells."
    )
    dash["E15"].alignment = Alignment(wrap_text=True, vertical="top")
    dash["E15"].font = font(size=11)

    # Helper columns for a simple chart of money
    dash["H6"] = "Money snapshot"
    dash["H6"].font = Font(name="Calibri", size=14, bold=True, color=GOLD)
    dash["H6"].fill = fill(NAVY)
    dash["I6"] = ""
    dash["I6"].fill = fill(NAVY)
    dash["H7"] = "Committed"
    dash["I7"] = "=C7"
    dash["H8"] = "Received"
    dash["I8"] = "=C8"
    dash["H9"] = "Remaining"
    dash["I9"] = "=C9"
    for r in (7, 8, 9):
        dash.cell(r, 9).number_format = '"$"#,##0'
        dash.cell(r, 9).fill = fill(GRAY)
        dash.cell(r, 9).border = thin

    chart = BarChart()
    chart.type = "col"
    chart.title = None
    chart.y_axis.title = None
    chart.x_axis.title = None
    chart.style = 10
    chart.y_axis.numFmt = "$#,##0"
    chart.legend = None
    data = Reference(dash, min_col=9, min_row=6, max_row=9)
    cats = Reference(dash, min_col=8, min_row=7, max_row=9)
    chart.add_data(data, from_rows=False, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 12
    chart.height = 6
    dash.add_chart(chart, "B19")

    dash["B34"] = "Open the Sponsors or Attendance tabs to type. Totals on this page follow those tables."
    dash["B34"].font = font(size=11, color="6B5B45")
    dash.merge_cells("B34:F34")

    # Print area
    dash.print_title_rows = "1:2"
    dash.oddHeader.left.text = "&K1C140E Artcell Edmonton Show"
    dash.oddFooter.left.text = "Yellow = type   ·   Gray = auto total"

    # --- How to ---
    howto.sheet_view.showGridLines = False
    apply_widths(howto, {"A": 3, "B": 88})
    howto["B1"] = "Put this file on your Microsoft 365 account"
    howto["B1"].font = Font(name="Calibri", size=20, bold=True, color=NAVY)
    howto.merge_cells("B1:B2")
    steps = [
        "1. Open OneDrive or SharePoint in the browser (onedrive.live.com or your work/school 365).",
        "2. Upload Artcell-Edmonton-Show.xlsx. Or File → Save As → OneDrive from Excel.",
        "3. Click Share → Anyone with the link can edit (or your org only, if that’s safer).",
        "4. Paste that link in the group chat. People tap it on their phone and choose Open in Excel.",
        "5. On a phone: install the Excel app if the browser feels cramped. Yellow cells are the only ones to type.",
        "",
        "Money: set the target on Dashboard, then enter Committed / Received on Sponsors. Remaining calculates.",
        "Attendance: add each person or group, how many seats, and status. Confirmed seats roll up to Dashboard.",
        "Setlist: tap a YouTube link to jump to that cue.",
        "",
        "Excel is simpler for totals. The phone website is simpler for big tap buttons. Use both if you want — they are not linked yet.",
        "Don’t copy this file around as email attachments after you start sharing. One file in OneDrive is the source of truth.",
    ]
    for i, line in enumerate(steps, start=4):
        howto.cell(i, 2, line).font = font(size=13)
        howto.row_dimensions[i].height = 22

    out = ROOT / "public" / "Artcell-Edmonton-Show.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    # also keep a copy at repo root for easy find
    root_copy = ROOT / "Artcell-Edmonton-Show.xlsx"
    wb.save(root_copy)
    return out


if __name__ == "__main__":
    path = build()
    print(path)
